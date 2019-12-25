# -*- coding: utf-8 -*-
from scripts.handle_mysql import *
import requests
from scripts.handle_excel import *
import unittest
from openpyxl import load_workbook
from scripts.handle_log import do_logger
import requests,json,unittest,os,sys
import unittest
import json, os, sys
import datetime
print(os.path.split(os.path.dirname(__file__))[0])
file_name = os.path.join((os.path.split(os.path.dirname(__file__))[0]),'datas/case2.xlsx')
#sys.path.append(aa)
print(file_name)
class Test_首页帖子(unittest.TestCase):
    dict_showtype = {"0": "未知,错误,无广告,头条SDK 广告 ",
                     "1": "开屏图片广告", "22": "开屏视频广告", "2": "信息流大图广告 ", "3": "信息流小图广告",
                     "4": "信息流三图广告 ", "6": "信息流视频广告", "8": "内页悬浮logo广告",
                     "13": "新闻背景广告 ", "16": "下拉海报图片广告 ", "29": "下拉海报图片广告(动图)",
                     "17": "下拉海报视频广告(1 / 3", "21": "下拉海报视频广告(全屏)", "23": "帖子类型广告(纯文本)",
                     "27": "激励视频", "30": "更多页横幅大图广告", "31": "更多页横幅动图广告",
                     "32": "比赛推荐页信息流广告", "33": "搜索横幅大图广告", "34": "搜索横幅动图广告" }
    def __init__(self,*args,**kwargs):
        unittest.TestCase.__init__(self,*args,**kwargs)
        # def __init__(self, file_name=r"/Users/wangna/Downloads/API_NEWTest/datas/case2.xlsx", Sheet_name="首页帖子"):
        self.name = file_name
        self.Sheet_name = "关注"
        # self.real_path = os.path.join(path_data, self.name)
        self.data = load_workbook(file_name)
        self.sheet =self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.username = list(self.sheet["B"])
        self.pwd = list(self.sheet["C"])
        self.params = list(self.sheet["D"])
        self.vcode = list(self.sheet["E"])
        self.ex_res = list(self.sheet["F"])
        self.re_res = list(self.sheet["G"])
        self.is_pass = list(self.sheet["H"])
        self.nrows = self.sheet.max_row
    def test_社区关注(self):    #判断素材 materid
        # file_name = r"/Users/wangna/Downloads/API_NEWTest/datas/case2.xlsx"  #要变成相对路径
        one_excel = HandleExcel(file_name,"关注")
        values = one_excel.get_cases()
        # print(values)
        import datetime
        t_today = datetime.datetime.now().strftime("%Y-%m-%d")
        print(type(t_today))
        url = "http://goblin.hupu.com/3/7.3.30/interfaceAd/getOther"
        print(t_today)
        for i in range(0,6):  #取出      self.nrows
            response = requests.request("GET", url,params=values[i]['params'])
            # if ress['ad']['id']:
            ress=response.json()
            do_logger.info("\n{:=^40s}".format("ress"))
            # print("返回值到数据类型是："+"type(ress)")
            # print("接口返回adcode值结果"+str(ress['ad_code']))
            adpid = values[i]['adpid']
            # adpid=ress['share_data']['app-request']['adpid']
            if ress['ad_code'] ==1 and ress["show_type"] !=0:  # m没有dsp怎么写
                # if ress['show_type']==6:  #信息流视频 video是数组   开机放两个视频 判断in关系
                #     print("这是adcc广告，要assert 素材id和link是否正确 -------------------" )
                # print(ress)
                ad_code =ress['ad_code']
                res_sucai_id = ress['ad']['id']
                self.sheet[('F' + str(i + 1))] = res_sucai_id
                self.data.save(file_name)
                self.sheet[('G' + str(i + 1))] = ad_code
                self.data.save(file_name)
                hupu_ad_type = ress['hupu_ad_type']
                # lp = ress['ad']['lp']
                # brand_name = ress['ad']['brand_name']
                # title = ress['ad']['title']
                # img = ress['ad']['img']  #list类型
                # video_url = ress['ad']['video_url']
                print('hupu_ad_type:'+hupu_ad_type+'接口返回素材id'+res_sucai_id)
                do_mysql = HandleMysql()
                result=do_mysql.run("select * from `cc_schedule_material` where `date` = '%s' and `adpid` = %s " % (t_today,adpid),is_more=True )
                print("-------sql----查出对应的素材id---")
                """
                取出来的数据放到对象里面
                判断purpose =1 showtimes =999 那么就去走下面的
                """
                print(result)
                sq_materid=result[0]['material_id']
                r_sid = result[0]['sid']
                print(result[0]['material_id'])
                result = do_mysql.run("select * from `cc_schedule_material` where `date` = '%s' and `adpid` = %s " % (t_today,adpid), is_more=True)
                print("-------sql----purpose---")
                print(result)
                sq_sid=result[0]['sid']
                print(result[0]['sid'])
                showtime=int(result[0]['show_times'])
                sq_purpose=do_mysql.run("select * from `cc_schedule` where `id` = '%s'  " % (sq_sid),is_more=True )
                # self.assertEqual(video_url, sq_video)
                purpose=int(sq_purpose[0]['purpose'])
                print(sq_purpose[0]['purpose'])
                self.assertEqual(res_sucai_id, str(sq_materid))
                print(values[i]['casename'] + "是品牌广告，接口返回素材id是" + str(res_sucai_id),"数据库的素材id："+str(sq_materid)+"校验通过")
                if purpose and showtime==999:
                    if purpose==4:
                        print(values[i]['casename'] + "purpose==4是内部投放")
                    else:
                        self.assertEqual(res_sucai_id,str(sq_materid))
                        self.sheet[('E' + str(i + 1))] = sq_materid
                        self.data.save(file_name)
                        print(values[i]['casename']+"是品牌广告，素材id是"+str(res_sucai_id))
                        pass_count=+1

            elif ress['ad_code'] !=1:
                if ress['ad_code']== 0:
                    print(values[i]['casename']+"不是品牌广告，返回adcode值是："+str(ad_code))
                else:
                    dsp = ress['ad']['dsp']
                    print(dsp)
                    ad_code = ress['ad_code']
                    do_logger.info("\n{:=^40s}".format("ress"))
                    self.sheet[('G' + str(i + 1))] = ad_code
                    self.data.save(file_name)
                    self.sheet[('H' + str(i + 1))] = dsp
                    self.data.save(file_name)
                    # print(ress['share_data']['placement-info']['name']+"不是品牌广告，返回adcode值是："+str(ad_code)+"dsp:"+str(dsp))
                    print(values[i]['casename'] + "不是品牌广告，返回adcode值是：" + str(ad_code) + ",dsp:" + str(dsp)+"校验后")

            elif ress['ad_code']==1 and ress["ad"]["dsp"]:
                # print("这是第三方")
                dsp = ress['ad']['dsp']
                ad_code = ress['ad_code']
                do_logger.info("\n{:=^40s}".format("ress"))
                self.sheet[('G' + str(i + 1))] = ad_code
                # self.sheet[('E' + str(i + 1))] = materid
                self.data.save(file_name)
                print(values[i]['casename'] + "不是品牌广告，返回adcode值是：" + str(ad_code) + ",dsp:" + str(dsp))

            elif ress['ad']:
                if ress['ad']['dsp']:
                    dsp = ress['ad']['dsp']
                    ad_code =ress['ad_code']
                    do_logger.info("\n{:=^40s}".format("ress"))
                    self.sheet[('G' + str(i + 1))] = ad_code
                    # self.sheet[('E' + str(i + 1))] = materid
                    self.data.save(file_name)
                    print(values[i]['casename'] + "不是品牌广告，返回adcode值是：" + str(ad_code) + ",dsp:" + str(dsp))
                else:
                   print(values[i]['casename'] + "不是品牌广告，返回adcode值是：" + str(ad_code))


if __name__ == '__main__':
    unittest.main()


#吧查出来的数据写入到excel内
