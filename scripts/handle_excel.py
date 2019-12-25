# -*- coding: utf-8 -*-
"""
-------------------------------------------------
  @Time : 2019/6/25 22:03 
  @Auth : 可优
  @File : handle_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
-------------------------------------------------
"""
from openpyxl import load_workbook
from scripts.handle_config import do_config
class HandleExcel:
    """
    定义处理excel的类
    """
    def __init__(self, filename, sheetname=None):
        self.filename, self.sheetname = filename, sheetname
    def get_cases(self):
        # 1. 打开excel
        wb = load_workbook(self.filename)   # 相当于一个excel文件
        # 2.定位表单
        # 如果你有传sheetname, 那么就获取指定的表单
        # 如果你没有传sheetname, 那么就获取第一个表单
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        # 获取所有的用例数据
        # 如果values_only为False(默认), 那么返回的是cell对象组成的生成器(需要将其转化为列表或者元祖)
        # 如果values_only为True, 那么返回的是单元格中的值组成的生成器(需要将其转化为列表或者元祖)
        # one_value = ws.iter_rows(min_row=2, max_row=5, values_only=True)
        values = tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        sheet_head_tuple = values[0]

        # print(sheet_head_tuple)
        # value_tuple = tuple( one_value)  # 每一行单元格组成的, 嵌套元祖的元祖
        cases_list = []  # 将用例字典信息, 存放在列表中, 这个列表就是嵌套字典列表
        for data in tuple(ws.iter_rows(min_row=2, values_only=True)):
            cases_list.append(dict(zip(sheet_head_tuple, data)))

        return cases_list

    def write_result(self, row, actual, result):
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]

        # if isinstance(row, int) and (2 <= row <= other_ws.max_row):
        if isinstance(row, int) and (row >= 2):
            other_ws.cell(row=row,
                          column=do_config.get_int('excel', 'actual_col'),
                          value=actual)
            other_ws.cell(row=row,
                          column=do_config.get_int('excel', 'result_col'),
                          value=result)
            other_wb.save(self.filename)
        else:
            print("传入的行号有误,行号应大于1的整数")


if __name__ == '__main__':
    # file_name = "/Users/wangna/Downloads/API_NEWTest/datas/case2.xlsx"
    # one_excel = HandleExcel(file_name)
    # print(one_excel)
    # values = one_excel.get_cases()
    # print(values)
   # values one_excel.write_result(3, 100, "KeYou")
    wb2 = load_workbook(r"/Users/wangna/Downloads/API_NEWTest/datas/case2.xlsx")
    ws = wb2["首页帖子"]
    print(ws)
    print(ws["params"][2].value)

